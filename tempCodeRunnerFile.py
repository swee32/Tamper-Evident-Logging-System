import pandas as pd
import mysql.connector
import hashlib
import config
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def export_secure_excel():
    try:
        # 1. Connect using restricted locker_admin credentials
        db = mysql.connector.connect(**config.DB_CONFIG)
        query = "SELECT * FROM access_logs ORDER BY id ASC"
        df = pd.read_sql(query, db)
        db.close()

        if df.empty:
            print("⚠️ No logs found in the database to export.")
            return

        # 2. Save initial Excel file
        file_name = "Security_Audit_Report.xlsx"
        df.to_excel(file_name, index=False)

        # 3. Open with Openpyxl for Forensic Formatting
        wb = load_workbook(file_name)
        ws = wb.active

        # Define your preferred alert colors
        pink_fill = PatternFill(start_color="#FFC7CE", end_color="#FFC7CE", fill_type="solid") # Light Pink
        red_fill = PatternFill(start_color="#FF9999", end_color="#FF9999", fill_type="solid")  # Light Red

        chain_broken = False
        last_hash = "0" * 64 # Start with Genesis hash

        # Iterate through rows (Excel rows start at 2 because of header)
        for i, row in enumerate(df.itertuples(), start=2):
            # Check A: Chain Integrity (Is the chain broken?)
            if chain_broken or row.prev_hash != last_hash:
                chain_broken = True
                # Highlight entire row Light Pink
                for cell in ws[i]:
                    cell.fill = pink_fill   # Highlight THIS and ALL subsequent rows
            
            if not chain_broken:
                last_hash = row.current_hash
            # If chain broken, last_hash remains incorrect to maintain detection

            
            # Check B: Data Integrity (Has content been modified?)
            data_to_verify = f"{row.timestamp}{row.event_type}{row.description}{row.prev_hash}"
            recalculated_hash = hashlib.sha256(data_to_verify.encode()).hexdigest()

            if row.current_hash != recalculated_hash:
                # Highlight specific data cells Light Red
                ws.cell(row=i, column=2).fill = red_fill # Timestamp
                ws.cell(row=i, column=4).fill = red_fill # Description
                ws.cell(row=i, column=6).fill = red_fill # Current Hash
            
            last_hash = row.current_hash

        wb.save(file_name)
        print(f"✅ Secure Forensic Report Generated: {file_name}")

    except Exception as e:
        print(f"❌ Error during Excel export: {e}")

if __name__ == "__main__":
    export_secure_excel()