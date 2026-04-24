import pandas as pd
import mysql.connector
import hashlib
import config
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

def export_secure_excel():
    try:
        # 1. Connect using restricted locker_admin credentials
        db = mysql.connector.connect(**config.DB_CONFIG)
        query = "SELECT * FROM access_logs ORDER BY id ASC"
        df = pd.read_sql(query, db)
        db.close()

        if df.empty:
            print("⚠️ No logs found in the database to export.")
            return

        # 2. Save initial Excel file
        file_name = "Security_Audit_Report.xlsx"
        df.to_excel(file_name, index=False)

        # 3. Open with Openpyxl for Forensic Formatting
        wb = load_workbook(file_name)
        ws = wb.active

        # Define colors in aRGB format (matches report: #FF9999 and #FFC7CE)
        pink_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
        red_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
        
        # Define alignment with wrap text
        wrap_alignment = Alignment(
            horizontal='left',
            vertical='top',
            wrap_text=True,
            shrink_to_fit=False
        )
        
        # Center alignment for ID column
        center_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=False
        )

        # Apply alignment to ALL cells first (header + data)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = wrap_alignment
        
        # Special: Center align the ID column (column A)
        for row in ws.iter_rows(min_col=1, max_col=1):
            for cell in row:
                cell.alignment = center_alignment

        # Set column widths for better readability
        column_widths = {
            'A': 8,     # id
            'B': 20,    # timestamp
            'C': 18,    # event_type
            'D': 50,    # description
            'E': 70,    # prev_hash
            'F': 70,    # current_hash
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Set header row height
        ws.row_dimensions[1].height = 30

        chain_broken = False
        last_hash = "0" * 64

        # Iterate through rows (Excel rows start at 2 because of header)
        for i, row in enumerate(df.itertuples(), start=2):
            # Check A: Chain Integrity (Is the chain broken?)
            if chain_broken or row.prev_hash != last_hash:
                chain_broken = True
                # Highlight entire row Light Pink
                for cell in ws[i]:
                    cell.fill = pink_fill
            
            # Update last_hash ONLY if chain is intact
            if not chain_broken:
                last_hash = row.current_hash
            
            # Check B: Data Integrity (Has content been modified?)
            data_to_verify = f"{row.timestamp}{row.event_type}{row.description}{row.prev_hash}"
            recalculated_hash = hashlib.sha256(data_to_verify.encode()).hexdigest()

            if row.current_hash != recalculated_hash:
                # Highlight ALL data cells Light Red
                ws.cell(row=i, column=2).fill = red_fill  # Timestamp
                ws.cell(row=i, column=3).fill = red_fill  # Event Type
                ws.cell(row=i, column=4).fill = red_fill  # Description
                ws.cell(row=i, column=5).fill = red_fill  # Prev Hash
                ws.cell(row=i, column=6).fill = red_fill  # Current Hash
            
            # Auto-adjust row height based on content
            ws.row_dimensions[i].height = None

        # Freeze the header row
        ws.freeze_panes = 'A2'

        wb.save(file_name)
        print(f"✅ Secure Forensic Report Generated: {file_name}")

    except Exception as e:
        print(f"❌ Error during Excel export: {e}")

if __name__ == "__main__":
    export_secure_excel()